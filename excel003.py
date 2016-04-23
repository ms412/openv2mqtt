import os
import openpyxl

path = r'demo1.xlsx'

book = openpyxl.Workbook()

sheet = book.get_active_sheet()
sheet.title = 'Test'

sheet.cell('A1').value ='ff'
#cell_A1.value = 'a1'

sheet.cell( row=2, column=1 ).value = 'c1'

book.save(path)