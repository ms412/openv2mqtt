import os
import openpyxl

#------------------------------------------

# Change this to suit
path = r'demo.xlsx'

 # Create an empty workbook
book = openpyxl.Workbook()

# Get the only sheet
sheet = book.get_active_sheet()
sheet.title = 'Test'

# Write to individual cells
cell_A1 = sheet.cell( 'A1' )
cell_A1.value = 'a1'

# the row and column assignments are needed
cell_B1 = sheet.cell( row=0, column=1 )
cell_B1.value = 'b1'

# Write some numerical data
data = (1.1,200,3+7j)
rng = sheet.range('A5:C5')
for r in rng:
    for cell, d  in zip( r, data ):
        cell.value = '9'

book.save(path)