import os.path
import openpyxl
import json



class excel(object):

    def __init__(self, config,msg):

        self._file = config.get('FILE')
        self._sheetname = config.get('SHEETNAME','Sheet1')
        #Worbook Handle
        self._wbHandle = None
        #Sheet Handle
        self._sheetHandle = None
        #Headline
        self._headline = []

    def start(self):

        if self.excelFileExist():
            self.excelOpenWorkbook()
            self.excelOpenSheet()
            print(self.excelReadRow(2,4))
            #self.excelGetHaeadline()

        else:
            self.excelNewWorkbook()
            self.excelNewSheet()
            self.excelSaveFile()

    def excelFileExist(self, filename = None):
        '''
        checks if file exist
        :return:
        True if file exist
        False if file does not exist
        '''
        result = False
        if filename == None:
            filename = self._file

        if os.path.isfile(filename):
            result = True

        return result

    def excelSaveFile(self, filename = None):

        if filename == None:
            filename = self._file

        self._wbHandle.save(filename)

    def excelOpenWorkbook(self,filename = None):
        if filename == None:
            filename = self._file
        self._wbHandle = openpyxl.load_workbook(filename,use_iterators=True)
        return True

    def excelNewWorkbook(self):
         # Create an empty workbook
        self._wbHandle = openpyxl.Workbook()
        return True

    def excelNewSheet(self,name=None):
        self._sheetHandle = self._wbHandle.get_active_sheet()

        if name == None:
            name =  self._sheetname

        self._sheetHandle.title = name

    def excelOpenSheet(self, name=None):
        if name == None:
            name =  self._sheetname

        self._sheetHandle = self._wbHandle.get_sheet_by_name(name)

    def excelReadRow(self,start, stop = None):
        '''
        reads one or more rows
        if no stop parameter is given, only one row will written
        :param start: start row 1...n
        :param stop:
        :return: list of rows
        '''
        result = []

        if stop == None:
            stop = start

        start = (start-1)

        print(start,stop)
        for index, row in enumerate(self._sheetHandle.iter_rows()):
            print(index)
            if start <= index and stop > index:
                print('Row',row)
                result.append(row)
                for cell in row:
                    print(cell.coordinate,cell.value)

        return result

    def excelWriteRow(self,data,rowId=None):

        if rowId == None:
            rowId = 1

        wb2 = openpyxl.load_workbook('demo1.xlsx')
        ws4 = wb2.get_sheet_by_name("Test")

        ws4.cell(row=7,column=2).value='f'

        #x.value='f'

        for index, item in enumerate(data, start=1):
            print(rowId,index,item)
            ws4.cell(row=rowId, column=index).value = item
           # ws.cell(row=i+2, column=0).value = statN

        wb2.save('demo1.xlsx')

    def excelGetHaeadline(self):

        for index, row in enumerate( ):
            print('Row:', index, row)

       # for row in self._sheetHandle.iter_rows():
        #    print(row)
         #   for cell in row:
          #      print (cell.coordinate, cell.value)
'''
        for item in self._sheetHandle.rows:
            print(item.value)
            self._headline.append(item)

        print('Headline',self._headline)
'''
       # self._sheetHandle.get_highest_row()
        #print(self._sheetHandle.get_highest_row())
'''
        header_row = None
for i, row in enumerate(sheet.iter_rows()):
  if header_row == None:
    header_row = [ c.internal_value for c in row ]
    continue
  row = dict(zip_longest(header_row, [ c.internal_value for c in row ]))
  print(row)

'''




if __name__ == "__main__":

    j = """[{"getTempAtp": {"Type": "Grad Celsius", "VALUE": "11.200000"}, "getTempStp2": {"Type": "Grad Celsius", "VALUE": "20.000000"}, "getBrennerStarts": {"Type": "", "VALUE": "993.000000"}, "getPumpeStatusIntern": {"Type": "1", "VALUE": "1"}, "getWWUWPNachlauf": {"Type": "", "VALUE": "2.000000"}, "getTempRaumNorSollM1": {"Type": "Grad Celsius", "VALUE": "18.000000"}, "getTempKist": {"Type": "Grad Celsius", "VALUE": "36.500000"}, "getTempVListM1": {"Type": "Grad Celsius", "VALUE": "36.500000"}, "getTempVListM2": {"Type": "Grad Celsius", "VALUE": "23.000000"}, "getPumpeStatusM1": {"Type": "%", "VALUE": "45.000000"}, "getTempWWist": {"Type": "Grad Celsius", "VALUE": "47.000000"}, "getTempA": {"Type": "Grad Celsius", "VALUE": "11.300000"}, "getSpeichervorrang": {"Type": "", "VALUE": "2.000000"}, "getTempAbgas": {"Type": "Grad Celsius", "VALUE": "36.000000"}, "ngetTempVLsollM2": {"Type": "command unknown", "VALUE": "ERR:"}, "getPumpeStatusM2": {"Type": "OK", "VALUE": "NOT"}, "getTempKsoll": {"Type": "Grad Celsius", "VALUE": "28.900000"}, "getStatusStoerung": {"Type": "0", "VALUE": "0"}, "getTempVLsollM1": {"Type": "Grad Celsius", "VALUE": "21.000000"}, "getPumpeStatusZirku": {"Type": "0", "VALUE": "0"}, "getPumpeDrehzahlIntern": {"Type": "%", "VALUE": "55.000000"}, "getVolStrom": {"Type": "l/h", "VALUE": "1017.000000"}, "getTempWWsoll": {"Type": "Grad Celsius", "VALUE": "45.000000"}, "getTempRL17A": {"Type": "Grad Celsius", "VALUE": "35.500000"}}]"""
    jmsg = json.loads(j)

    cfgdict = {}
    cfgdict['FILE']= 'exceltest001.xlsx'
    cfgdict['SHEETNAME'] = 'TEST123'
    test = ['uu','ii','oo','pp']

    ex = excel(cfgdict,jmsg)
   # ex.start()
    ex.excelWriteRow(test)
    ex.excelWriteRow(test,5)
  #  ex.excelSaveFile('exceltest001.xlsx')